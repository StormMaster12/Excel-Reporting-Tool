# -*- coding: utf-8 -*-

import openpyxl
import xml
import itertools
from datetime import date, datetime
from dateutil.parser import parse
import re
import sys
import glob
import logging
import logging.handlers
from operator import itemgetter
import codecs
import tkinter as tk


class Report_Generator():

    def Main(self):

        xml_iter,excel_workbook = self.Load()

        report = """<meta charset="UTF-8"> <html>
        <head>
        <!-- Plotly.js -->
        <script src="http://ajax.googleapis.com/ajax/libs/jquery/1.11.1/jquery.min.js"></script>
        <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
        <style>
        .expand1 { display: none;
        }

        .expand2 { display: none;
        }
        </style>

        <script> $(document).ready(function(){
            $(".btn1").click(function(){
                $(".expand1").toggle();});
            $(".btn2").click(function(){
                $(".expand2").toggle();});
        }) </script>

        
        </head>"""

        for node in xml_iter:
            if node.tag=="txt":
                report += "\n" + "<body><p>" + node.text + "</p></body>"
        
            elif node.tag=="SQL":

                SQL_Operators = ['SELECT','FROM','WHERE','AND','SUM','AVG','COUNT','MAX','MIN','GROUP BY']
                SQL_Maths_Operators = ['SUM','AVG','COUNT','MAX','MIN']
            
                SQL_Command_Dict = {}
                Output_List = []
                Group_By = []

                self.parse_SQL(node.text,SQL_Command_Dict,SQL_Operators,SQL_Maths_Operators)
                active_worksheet = excel_workbook[SQL_Command_Dict['FROM'][0]]
                self.find_columns(active_worksheet,SQL_Command_Dict)
                self.SQL_Logic(active_worksheet,SQL_Command_Dict,Group_By,Output_List)
                output = self.output_string_manipulation(SQL_Command_Dict,Output_List,Group_By)

                report +="\n" + "<body padding:0;margin:0;>" + output  + "</body>"

        report += "</html>"
        html_output = codecs.open("report.html","w","utf-8")
        html_output.write(report)
        html_output.close()

    def Load(self):
        my_logger.info('Program Has Started')
        #Loads the xml file. Controls the layout of the html

        self.file_locations = []
        self.tk_window()

        try:
            xml_tree = xml.etree.ElementTree.parse(self.file_locations[0])
            xml_root = xml_tree.getroot()
            xml_iter = xml_tree.iter()

            #Loads the excel file into memory
            excel_workbook = openpyxl.load_workbook(self.file_locations[1],data_only=True)
        except Exception as e:
            my_logger.error("Failed to Open File",exc_info=True)
            return;

        return xml_iter,excel_workbook

    def parse_SQL(self,SQL,SQL_Command_Dict,SQL_Operators,SQL_Maths_Operators):

        self.get_parsed_SQL(SQL,SQL_Command_Dict,SQL_Operators)

        SQL_Group_By =[]
        SQL_Select = []
        SQL_Math = []
        SQL_Where = []
        SQL_New_Math = []

        current_key = 'GROUP BY'
        try:
            if len(SQL_Command_Dict[current_key]) > 0:
                for sublist in SQL_Command_Dict['GROUP BY']:
                        a = sublist.replace(",","")
                        SQL_Group_By.append([a])
                SQL_Command_Dict[current_key] = SQL_Group_By
        except KeyError as e:
            print(False)

        current_key = 'SELECT'
        self.get_comma_list(SQL_Select,SQL_Command_Dict[current_key])


        current_key = 'AND'
        self.parse_And(SQL_Command_Dict[current_key],SQL_Where)

        #Maths Selectors      
        self.parse_Maths(SQL_Maths_Operators,SQL_Command_Dict,SQL_Math)

        SQL_Command_Dict['SELECT'] = SQL_Select
        SQL_Command_Dict['WHERE'] = SQL_Where
        SQL_Command_Dict['MATHS'] = SQL_Math

        del SQL_Select
        del SQL_Where
        del SQL_Math
        del SQL_Command_Dict['AND']

        if SQL_Command_Dict['MATHS']  and  not SQL_Command_Dict['GROUP BY']:
            my_logger.info('Group by Required When Maths operators present') 
            raise NameError('Group by Required When Maths operators present') 

    def find_columns(self,active_worksheet,SQL_Command_Dict):

        current_key = 'SELECT'
        self.get_column(active_worksheet,SQL_Command_Dict[current_key],current_key,SQL_Command_Dict)
        self.get_column(active_worksheet,SQL_Command_Dict[current_key],'WHERE',SQL_Command_Dict)
        self.get_column(active_worksheet,SQL_Command_Dict[current_key],'GROUP BY',SQL_Command_Dict)
        current_key = 'MATHS'
        self.get_column(active_worksheet,SQL_Command_Dict[current_key],current_key,SQL_Command_Dict)

    def SQL_Logic(self,active_worksheet,SQL_Command_Dict,Group_By,Output_List):



        #Logic Bit of the code
        #First loop goes through all of the rows in the spreadsheet

        for i in range(1,active_worksheet.max_row):
            count = 0
            #This goes through the commands given by the SQL
            for sublist in SQL_Command_Dict['WHERE']:
                sheet_data = active_worksheet.cell(row=i,column=sublist[-1]).value
                #Checks to see if the cell value is a datetime or not. Following operations can only be done on datetime objects
                if sublist[1] == 'BETWEEN':
                    if type(sheet_data) is datetime and self.validate_date(sublist[2]):
                            date1 = sheet_data.date()
                            #Converts the two dates provided by SQL_Where into datetime objects so they can be compared to the cell datetime
                            str_to_date1 =  self.to_date(sublist[2])
                            str_to_date2 = self.to_date(sublist[3])
                            if str_to_date1<date1 and date1<str_to_date2:
                                #If matches the conditions adds 1 to the count so that the final check knows if all checks have passed for a row.
                                count += 1
                elif sublist[1]== '=':
                    if sheet_data == sublist[2]:
                        count +=1

            #If all the checks have passed, the counter will equal the total number of clauses given. Will then add data from the column specified by SQL_Sun_Index
            if count == len(SQL_Command_Dict['WHERE']):
                self.return_table_data(SQL_Command_Dict,active_worksheet,i,Group_By,Output_List)

        del SQL_Command_Dict['WHERE'],active_worksheet

    def output_string_manipulation(self,SQL_Command_Dict,Output_List,Group_By):

        Output_String = ''

        Output_List.sort(key=itemgetter(0))
        Group_By.sort(key=itemgetter(0))

        header_list_group = []
        header_list_select = []

        Group_By_Output_List = []

        self.graph_output = ""
        self.return_group_by(Group_By,Group_By_Output_List,header_list_group,SQL_Command_Dict['MATHS'])

        Output_String += '\n <button type="button" class="btn1">Toggle the Table Below</button>'

        Output_String += '<div class=expand1>'

        Output_String += '\n'.join(self.generate_html_table(Output_List,SQL_Command_Dict['SELECT'],header_list_select,1))
        Output_String += '</div>'

        del Output_List, header_list_select,SQL_Command_Dict['SELECT']

        Output_String += '\n <button type="button" class="btn2">Toggle the Table Below</button>'
        Output_String += '<div class=expand2>'
        Output_String += '\n'.join(self.generate_html_table(Group_By_Output_List,SQL_Command_Dict['GROUP BY'],header_list_group,2))
        Output_String += '</div>'

        del Group_By_Output_List,SQL_Command_Dict['GROUP BY'],header_list_group

        Output_String += self.graph_output

        return  Output_String

    def return_table_data(self,SQL_Command_Dict,active_worksheet,i,Group_By,Output_List):

        flat_list = []
        Group_By_Flat_List = []

        for select_sublist in SQL_Command_Dict['SELECT']:
            sheet_value = active_worksheet.cell(row=i,column=select_sublist[-1]).value
            if type(sheet_value) is datetime:
                flat_list.append(datetime.strftime(sheet_value,'%d-%m-%Y'))
            else:
                flat_list.append(sheet_value)

        for group_by_sublist in SQL_Command_Dict['GROUP BY']:
            if type(active_worksheet.cell(row=i,column=group_by_sublist[-1]).value) is datetime:
                Group_By_Flat_List.append(datetime.strftime(active_worksheet.cell(row=i,column=group_by_sublist[-1]).value,'%d-%m-%Y'))
            else:
                Group_By_Flat_List.append(active_worksheet.cell(row=i,column=group_by_sublist[-1]).value)

        for maths_sublist in SQL_Command_Dict['MATHS']:
            flat_list.append('£'+ str(active_worksheet.cell(row=i,column=maths_sublist[-1]).value))
            Group_By_Flat_List.append(active_worksheet.cell(row=i,column=maths_sublist[-1]).value)

        Group_By.append(Group_By_Flat_List)
        Output_List.append(flat_list)

    def get_parsed_SQL(self,SQL,SQL_Command_Dict,SQL_Operators):

        SQL_Command_Pos = [] 
        SQL_No_Line = SQL.replace("\n"," ").replace('\r',' ') + " "

        for string in SQL_Operators:
            for m in re.finditer(string, SQL_No_Line):
                SQL_Command_Pos.append([m.start(),m.end(),string])
        SQL_Command_Pos.sort()

        for i in range(0,len(SQL_Command_Pos)): SQL_Command_Dict[SQL_Command_Pos[i][2]] = []

        for i in range(0,len(SQL_Command_Pos)):
            pm=''
            count=0         
            for m in re.finditer(' ', SQL_No_Line):
                if i < (len(SQL_Command_Pos)-1):
                    if m.start() >= SQL_Command_Pos[i][1] and m.end()<=SQL_Command_Pos[i+1][0]:
                        chunk_string = " ".join(self.get_string(SQL_No_Line,count,m,pm,SQL_Command_Pos[i][1]).split())
                        if chunk_string != SQL_Command_Pos[i][2]:
                            SQL_Command_Dict[SQL_Command_Pos[i][2]].append(self.get_string(SQL_No_Line,count,m,pm,SQL_Command_Pos[i][1]))
                elif i == (len(SQL_Command_Pos)-1):
                    if m.start()> SQL_Command_Pos[i][1]:                     
                        chunk_string = " ".join( self.get_string(SQL_No_Line,count,m,pm,SQL_Command_Pos[i][1]).split())
                        if chunk_string != SQL_Command_Pos[i][2]:
                            SQL_Command_Dict[SQL_Command_Pos[i][2]].append(self.get_string(SQL_No_Line,count,m,pm,SQL_Command_Pos[i][1]))
                    count = 0
                pm=m
                count +=1

    def is_in(self,i_str,i_list):
        if i_str not in i_list:
            i_list.append(i_str)        

    def return_maths(self,SQL_Command_Dict,Min_Max_Count,m_list,header_list):
        for sublist in SQL_Command_Dict:
            if 'MAX' in sublist:
                m_list.append( str(max(Min_Max_Count)))
                self.is_in('MAX',header_list)
            if 'MIN' in sublist: 
                m_list.append(str(min(Min_Max_Count)))
                self.is_in('MIN',header_list)
            if 'AVG' in sublist: 
                m_list.append((sum(Min_Max_Count)/len(Min_Max_Count)))
                self.is_in('AVG',header_list)
            if 'SUM' in sublist: 
                m_list.append(str(sum(Min_Max_Count)))
                self.is_in('SUM',header_list)
            if 'COUNT' in sublist: 
                m_list.append(str(len(Min_Max_Count)))
                self.is_in('COUNT',header_list)

    def get_comma_list(self,list_dict,dict_arr):

        for i in range(0,len(dict_arr)):
            if "," in dict_arr[i] or i == len(dict_arr)-1:
                list_dict.append([dict_arr[i-1].replace("_"," "), dict_arr[i].strip(",")])

    def parse_Maths(self,Operators_Dict,Main_dict,export_list):

        for i in range(0,len(Operators_Dict)):
            if Operators_Dict[i] in Main_dict.keys():
                for j in  range(0,len(Main_dict[Operators_Dict[i]])):
                    export_list.append([Main_dict[Operators_Dict[i]][j].strip('()'),Operators_Dict[i]])
        to_be_removed = []

        for i in range(0,len(export_list)):
            if i != len(export_list)-1:
                if export_list[i][0] in export_list[i+1]:
                    export_list[i].append(export_list[i+1][1])
                    to_be_removed.append(i+1)

        for i in range(0,len(to_be_removed)):
            del export_list[to_be_removed[i]-i]

    def parse_And(self,dict_arr,export_list):

        for i in range(0,len(dict_arr)):
            if dict_arr[i] == '=':
                if dict_arr[i+1][0] == "'" and dict_arr[i+1][-1] == "'" :
                    export_list.append([dict_arr[i-1],dict_arr[i],dict_arr[i+1].strip("'")])
                else:
                    contains = True
                    if "'" in dict_arr[i+1] and dict_arr[i+1].index("'") ==0 :
                        export_list.append([dict_arr[i-1],dict_arr[i],' '.join([dict_arr[i+1],dict_arr[i+2]]).strip("'")])
            elif  dict_arr[i] == "BETWEEN":
                export_list.append([ dict_arr[i-1],dict_arr[i], dict_arr[i+1], dict_arr[i+2]])

    def to_date(self,str_to_date): return datetime.strptime(str_to_date, '%d/%m/%Y').date()

    def get_string(self,string,count,m,pm,SQL_Command_Pos):
        if count == 0:
            return (string[SQL_Command_Pos:m.start()])
        else:
            return (string[pm.end():m.start()])

    def validate_date(self,test):
        try:
           datetime.strptime(test,'%d/%m/%Y')
        except ValueError:
            return False
        else:
            return True

    def get_column(self,active_worksheet,current_key,other_key,SQL_Dict):
        
        for i in range(1,13):
            for dict_list in current_key:
                if active_worksheet.cell(row=1,column=i).value == dict_list[0]:
                    for sublist in SQL_Dict[other_key]:
                        if dict_list[0] in sublist or dict_list[1] in sublist:
                            if other_key == 'GROUP BY':
                                sublist.insert(0,dict_list[0])
                            sublist.append(i)

    def generate_html_table(self,Output_List,dict_arr,header_list,current_table):

        Output_String =""
        yield'<table padding:0;margin:0;> \n  <tr> \n'
        for substring in dict_arr:
            yield '   <th>' + str(substring[1]) + '</th>'

        for h_string in header_list:
            yield  '   <th>' + str(h_string) + '</th>'
        yield '  </tr>'

        for sublist in Output_List:
            yield '\n  <tr> \n'
            for l_string in sublist:
                yield'    <td >'+ str(l_string) + '</td>'
            yield '\n  </tr> \n'

        yield '</table>'

    def remove_same(self,list):
        checked_list = []
        for sublist in list:
            for i in range(0,len(sublist)):
                if sublist[i] not in checked_list or "£" in sublist[i]:
                    checked_list.append(sublist[i])
                else:
                    sublist[i] = ""
        del checked_list

    def return_group_by(self,Group_By,Group_By_Output_List,header_list_group,SQL_Command_Dict):
        Grouped_Items = []

        for y,items in itertools.groupby(Group_By,itemgetter(0)):
            Grouped_Items.append(list(items))

        Full_list = []
        
        count = 0
        for g_sublist in Grouped_Items:
            
            title_list = []
            flat_list = []
            int_list = [[] for _ in range(0,len(SQL_Command_Dict))]
            for i,sub_sublist in itertools.product(range(0,len(int_list)),g_sublist):
                for value in sub_sublist:
                    if type(value) is int or type(value) is float :
                        int_list[i].append(value)
                    else:
                        if value not in flat_list:
                            flat_list.append(value)
                            title_list.append(value)
            
            for item in int_list:
                self.return_maths(SQL_Command_Dict,item,flat_list,header_list_group)
                self.return_graph(item,'box',count,title_list)
            count += 1

            Group_By_Output_List.append(flat_list)

        del Grouped_Items

    def return_graph(self,int_list,type,count,title):
        self.graph_output += '<div id="graph' + str(count) + '"></div>'
        self.graph_output += '<script> var trace1 = { x:' + str(int_list) + ', \n type:"' + type + '''",\n name: "Set 1"};
        var data = [trace1]; \n var layout = {title: " ''' +str(title[0]) +'''"};
        Plotly.newPlot("graph''' + str(count) + '", data, layout) </script> \n'

    def tk_window(self):
        root = tk.Tk()
        button = tk.Button(root,text='Pick Xml File',command=lambda:  self.onbutton(button,button1,1,root))
        button1 = tk.Button(root,text='Pick Excel File',command=lambda: self.onbutton(button1,button,2,root))
        button.pack()
        button1.config(state='disabled')
        button1.pack()
        root.mainloop()

    def onbutton(self,button,button1,count,root):
        
        self.file_locations.append(tk.filedialog.askopenfilename())
        button.config(state='disabled')

        if count != 2:
            button1.config(state='active')
        else:
            root.destroy()

LOG_FILENAME = __name__

# Set up a specific logger with our desired output level
my_logger = logging.getLogger('MyLogger')
my_logger.setLevel(logging.DEBUG)



if __name__ == '__main__':
    Report_Generator().Main()
